# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook

from localVariables import skillList, raceClassDict, guildList, spellList

source_folder = "d:\Work\Coding\\battleboard_data_extraction\\battleboard-data-extraction\source_files\\"

fileList = ["Wulfric_baneguard_current_v_2021.1.xlsm",
            "emilie_oct_23.xlsm",
            ]

# test_workbook = load_workbook(source_folder + "Wulfric_baneguard_current_v_2021.1.xlsm")

# print(test_workbook.sheetnames)
# ['Instructions', 'Adventure Record', 'The Character', 'Magic', 'Power', 'Armour', 'Battleboard', 'Notes', 'Base', 'Variables', 'Tables', 'STD HQA']

# skill name maps to dict with 2 keys, 1 to track total ranks, 1 to track number of files that formed the total
# should look like
# skillMap = {'Bonebreak' : {'ranks' : 4,
#                            'characters' : 2},
#             'dodge'     : {'ranks' : 1,
#                            'characters' : 1}
#             } 

skillMap = {}

# spell name maps to dict with 4 leys, each to track the number of times bought by a character of a given archetype
# should look like
# spellMap = {'blindness' : {'warrior' : 0,
#                            'scout' : 2,
#                            'acolyte' : 1,
#                            'mage' : 4},
#             'lightning bolt' : {'warrior' : 0,
#                                 'scout' : 2,
#                                 'acolyte' : 1,
#                                 'mage' : 4},
#             } 

spellMap = {}

for file in fileList:

    currentWorkbook = load_workbook(source_folder + file)
    skillSheet = currentWorkbook['The Character']

    characterClass = raceClassDict[skillSheet.cell(row=2,column=2).value]['class']
    characterRace = raceClassDict[skillSheet.cell(row=2,column=2).value]['race']
    primaryGuild = guildList[(skillSheet.cell(row=3,column=2).value - 1)]

    # openpyxl starts at 1 for row and column
    for i in range(14,397):
        
        # allow for offset between row number on sheet and position in list
        skillName = skillList[(i-14)]
        rankValue = skillSheet.cell(row=i,column=3).value 

        if rankValue == None or rankValue == 0:
            continue
        else:
            if skillName not in skillMap:
                # skillMap[skillName] = {'ranks' : rankValue,
                #                        'characters' : 1}
                skillMap[skillName] = {'ranksAcolyte' : 0,
                                        'charactersAcolyte' : 0,
                                        'ranksMage' : 0,
                                        'charactersMage' : 0,
                                        'ranksScout' : 0,
                                        'charactersScout' : 0,
                                        'ranksWarrior' : 0,
                                        'charactersWarrior' : 0,}
            
            skillMap[skillName]['ranks' + characterClass] = skillMap[skillName]['ranks' + characterClass] + rankValue
            skillMap[skillName]['characters' + characterClass] = skillMap[skillName]['characters' + characterClass] + 1 

    spellSheet = currentWorkbook['Magic']

    for i in range(12,427):

        spellBought = spellSheet.cell(row=i,column=4).value
        # print(spellBought)
        # break
        spellName = spellList[(i-12)]

        if spellBought == None or spellBought == 0:
            continue
        else:
            if spellName not in spellMap:
                spellMap[spellName] = {'Warrior' : 0,
                                       'Scout' : 0,
                                       'Acolyte' : 0,
                                       'Mage' : 0}
            
            spellMap[spellName][characterClass] += 1

        

result = Workbook()
skillSheet = result.active
skillSheet.title = "Skill summary"

skillSheet.cell(row=1,column=1).value = 'Skill Name'
skillSheet.cell(row=1,column=2).value = 'Average of total'
skillSheet.cell(row=1,column=3).value = 'Total Count'
skillSheet.cell(row=1,column=4).value = 'Average of Acolyte'
skillSheet.cell(row=1,column=5).value = 'Acolyte Count'
skillSheet.cell(row=1,column=6).value = 'Average of Mage'
skillSheet.cell(row=1,column=7).value = 'Mage Count'
skillSheet.cell(row=1,column=8).value = 'Average of Scout'
skillSheet.cell(row=1,column=9).value = 'Scout Count'
skillSheet.cell(row=1,column=10).value = 'Average of Warrior'
skillSheet.cell(row=1,column=11).value = 'Warrior Count'


rowNum = 2
for skill in skillList:
    skillSheet.cell(row=rowNum, column=1).value = skill

    if skill in skillMap:

        acolyteRanks = skillMap[skill]['ranksAcolyte']
        acolyteNum = skillMap[skill]['charactersAcolyte']
        mageRanks = skillMap[skill]['ranksMage']
        mageNum = skillMap[skill]['charactersMage']
        scoutRanks = skillMap[skill]['ranksScout']
        scoutNum = skillMap[skill]['charactersScout']
        warriorRanks = skillMap[skill]['ranksWarrior']
        warriorNum = skillMap[skill]['charactersWarrior']

        totalRanks = acolyteRanks + mageRanks + scoutRanks + warriorRanks
        totalNum = acolyteNum + mageNum + scoutNum + warriorNum

        skillSheet.cell(row=rowNum, column=2).value = totalRanks / totalNum if totalNum != 0 else 0
        skillSheet.cell(row=rowNum, column=3).value = totalRanks
        skillSheet.cell(row=rowNum, column=4).value = acolyteRanks / acolyteNum if acolyteNum != 0 else 0
        skillSheet.cell(row=rowNum, column=5).value = acolyteNum
        skillSheet.cell(row=rowNum, column=6).value = mageRanks / mageNum if mageNum != 0 else 0
        skillSheet.cell(row=rowNum, column=7).value = mageNum
        skillSheet.cell(row=rowNum, column=8).value = scoutRanks / scoutNum if scoutNum != 0 else 0
        skillSheet.cell(row=rowNum, column=9).value = scoutNum
        skillSheet.cell(row=rowNum, column=10).value = warriorRanks / warriorNum if warriorNum != 0 else 0
        skillSheet.cell(row=rowNum, column=11).value = warriorNum
    else:
        skillSheet.cell(row=rowNum, column=2).value = 0
        skillSheet.cell(row=rowNum, column=3).value = 0
        skillSheet.cell(row=rowNum, column=4).value = 0
        skillSheet.cell(row=rowNum, column=5).value = 0
        skillSheet.cell(row=rowNum, column=6).value = 0
        skillSheet.cell(row=rowNum, column=7).value = 0
        skillSheet.cell(row=rowNum, column=8).value = 0
        skillSheet.cell(row=rowNum, column=9).value = 0
        skillSheet.cell(row=rowNum, column=10).value = 0
        skillSheet.cell(row=rowNum, column=11).value = 0

    rowNum += 1

spellSheet = result.create_sheet("Spells Summary")

spellSheet.cell(row=1,column=1).value = 'Spell Name'
spellSheet.cell(row=1,column=2).value = 'Times bought Total'
spellSheet.cell(row=1,column=3).value = 'Times bought Mage'
spellSheet.cell(row=1,column=4).value = 'Times bought Acolyte'
spellSheet.cell(row=1,column=5).value = 'Times bought Scout'
spellSheet.cell(row=1,column=6).value = 'Times bought Warrior'

rowNum = 2
for spell in spellList:
    spellSheet.cell(row=rowNum, column=1).value = spell

    if spell in spellMap:
        mageTotal = spellMap[spell]['Mage']
        acolyteTotal = spellMap[spell]['Acolyte']
        scoutTotal = spellMap[spell]['Scout']
        warriorTotal = spellMap[spell]['Warrior']
        
        total = mageTotal + acolyteTotal + scoutTotal + warriorTotal

        spellSheet.cell(row=rowNum, column=2).value = total
        spellSheet.cell(row=rowNum, column=3).value = mageTotal
        spellSheet.cell(row=rowNum, column=4).value = acolyteTotal
        spellSheet.cell(row=rowNum, column=5).value = scoutTotal
        spellSheet.cell(row=rowNum, column=6).value = warriorTotal
    else:
        spellSheet.cell(row=rowNum, column=2).value = 0
        spellSheet.cell(row=rowNum, column=3).value = 0
        spellSheet.cell(row=rowNum, column=4).value = 0
        spellSheet.cell(row=rowNum, column=5).value = 0
        spellSheet.cell(row=rowNum, column=6).value = 0

    rowNum += 1

result.save(source_folder + "results.xlsx")

# print(len(skillList))
# print(skillMap)
# print(spellList)
# print(spellMap)

